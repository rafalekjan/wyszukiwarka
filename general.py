import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time
from selenium import webdriver
import logging

now_date = time.strftime("%d.%m.%Y")
przetargi = []


class __NbrOfArt:
    def __init__(self, method_url, web_page_url, css_path):
        self.method_url = method_url
        self.web_page_url = web_page_url
        self.css_path = css_path

    def count_art(self):
        if self.method_url == "soup":
            self.web_page_get = requests.get(self.web_page_url)
            self.web_page_content = BeautifulSoup(self.web_page_get.content, 'html.parser')
            lp_art = len(self.web_page_content.select(self.css_path))
            return lp_art
        if self.method_url == "selenium":
            self.browser = webdriver.Chrome("C:\Python27\Scripts\chromedriver.exe")
            self.browser.implicitly_wait(7)
            self.browser.get(self.web_page_url)
            self.browser.find_elements_by_css_selector(self.css_path)
            lp_art = len(self.browser.find_elements_by_css_selector(self.css_path))
            return lp_art


class __DataOfArt(__NbrOfArt):
    def __init__(self, lp_art, filter_title, filter_url, base_url):
        self.title_of_art = []
        self.url_of_art = []
        self.data_of_art = []
        self.lp_art = lp_art
        self.filter_title = filter_title
        self.filter_url = filter_url
        self.base_url = base_url

    def data_title(self):
        for i in range(self.lp_art):
            if self.lp_art > 100 and i > 100:
                print("Przeanalizowano", i-1, "z", self.lp_art, "rekordów")
                log_file.write("Przeanalizowano" + str(i-1) + " z " + str(self.lp_art) + " rekordów" + "\n")
                self.lp_art = 100
                return self.title_of_art
            article_nth = obj__NbrOfArt.css_path + ":nth-of-type(" + str(i+1) + ")"
            try:
                if obj__NbrOfArt.method_url == "soup":
                    if self.filter_title is not None:
                        article_data = obj__NbrOfArt.web_page_content.select(article_nth)
                        article_title = article_data[0].select(self.filter_title)[0].text.lower()
                    else:
                        article_title = ""
                elif obj__NbrOfArt.method_url == "selenium":
                    if self.filter_title is not None:
                        article_title = obj__NbrOfArt.browser.find_element_by_css_selector(article_nth + ' > ' + self.filter_title).text.lower()
                    elif self.filter_title is None:
                        article_title = obj__NbrOfArt.browser.find_element_by_css_selector(article_nth).text.lower()
                    else:
                        article_title = ""
            except:
                print("~~!Błąd odczytu tytułu ogłoszenia", i+1)
                log_file.write("~~!Błąd odczytu tytułu ogłoszenia " + str(i+1) + "\n")
                article_title = ""
            self.title_of_art.append(article_title)
        return self.title_of_art

    def data_url(self):
        article_url = ""
        for i in range(self.lp_art):
            if self.title_of_art[i] == "":

                self.url_of_art.append(article_url)
                continue
            article_nth = obj__NbrOfArt.css_path + ":nth-of-type(" + str(i+1) + ")"
            try:
                if obj__NbrOfArt.method_url == "soup":
                    if self.filter_url is not None and self.filter_url != "base_url":
                        article_data = obj__NbrOfArt.web_page_content.select(article_nth)
                        for a in article_data[0].select(self.filter_url, href=True):
                            if self.base_url is not None:
                                article_url = str(self.base_url) + a['href']
                            else:
                                article_url = a['href']
                    else:
                        article_url = self.base_url
                elif obj__NbrOfArt.method_url == "selenium":
                    if self.filter_url is not None and self.filter_url != "base_url":
                        article_url = obj__NbrOfArt.browser.find_element_by_css_selector(
                            article_nth + ' > ' + self.filter_url).get_attribute('href')
                    elif self.filter_url is None and self.filter_url != "base_url":
                        article_url = obj__NbrOfArt.browser.find_element_by_css_selector(article_nth).get_attribute('href')
                    else:
                        article_url = self.base_url
            except:
                print("~~!Błąd odczytu adresu url ogłoszenia", i+1)
                log_file.write("~~!Błąd odczytu adresu url ogłoszenia " + str(i+1) + "\n")
                article_url = self.base_url
            self.url_of_art.append(article_url)
        return self.url_of_art

    def data_art(self):
        self.data_title()
        self.data_url()
        print("Znaleziono", self.lp_art, "ogloszen")
        for i in range(self.lp_art):
            if self.title_of_art[i] == "":
                continue
            else:
                self.data_of_art.append([self.title_of_art[i], self.url_of_art[i]])
        if obj__NbrOfArt.method_url == "selenium":
            obj__NbrOfArt.browser.quit()
        return self.data_of_art


class __ExcelHist:

    def __init__(self):
        self.new_auctions = []

    def write_history(self):
        for i in range(len(przetargi)):
            if len(history_list) == 0:
                self.new_auctions.append([przetargi[i][0], przetargi[i][1]])
            else:
                for history_record in range(len(history_list)):
                    if history_list[history_record] == przetargi[i][0]:
                        break
                    elif history_record == (len(history_list)-1):
                        self.new_auctions.append([przetargi[i][0], przetargi[i][1]])
                        break
        for i in range(len(self.new_auctions)):
            sheet_history.cell(row=len(history_list)+1, column=1).value = self.new_auctions[i][0]
            history_list.append(self.new_auctions[i][0])
            book_history.save('History\\' + book_words_list[customer]+'_history.xlsx')
        # print("Baza klienta", book_words_list[customer], "zawiera", len(history_list), "rekordów historycznych")
        log_file.write("Baza klienta " + str(book_words_list[customer]) + " zawiera " + str(len(history_list)) + " rekordów historycznych" + "\n")
        return self.new_auctions


class __ExcelArt:
    lp_rekordow_city = 0

    def __init__(self, row, customer):
        self.row = row
        self.customer = customer

    def write_excel(self):
        for i in range(len(new_auctions)):
            if filtr_slowo[self.customer][0][1] == 'tak':
                for word in range(len(filtr_slowo[self.customer][1])):
                    # print(filtr_slowo[self.customer][1][word])
                    if filtr_slowo[self.customer][1][word] in new_auctions[i][0]:
                        self.lp_rekordow_city += 1
                        self.row += 1
#######################
                        if len(new_auctions[i][1]) > 255:
                            sheet_results.cell(row=self.row, column=1).value = new_auctions[i][1]
                        else:
                            sheet_results.cell(row=self.row, column=1).value = '=HYPERLINK("{}", "{}")'.format(new_auctions[i][1], ">Link<")
                        if len(new_auctions[i][1]) < 5:
                            print("Link ogloszenia jest za krótki")
#######################
                        sheet_results.cell(row=self.row, column=2).value = ""
                        for a in range(int(len(new_auctions[i][0])/100+1)+1):
                            sheet_results.cell(row=self.row, column=2).value += new_auctions[i][0][a*100:(a+1)*100] + "\n"
                        sheet_results.row_dimensions[self.row].height = 80
                        break
        book_results.save('Results\\' + book_words_list[customer] + '_results.xlsx')
        return self.row

##################################################################################################################
log_file = open('Logs\\Log_file_' + now_date + '_' + 'Webpages.txt', 'a')
log_file.write(str(now_date) + "\n")
try:
    filtr_slowo = []
    book_words = load_workbook('control.xlsx', data_only=True)
    book_words_list = book_words.sheetnames
    book_words_list.remove('MAIN')
    for index in range(len(book_words_list)):
        sheet_words = book_words[book_words_list[index]]
        # filtr_slowo.append([sheet_words['G1'].value, sheet_words['G9'].value.lower(), sheet_words['A2'].value.lower(), sheet_words['A4'].value.lower()])
        filtr_slowo.append([[sheet_words['G1'].value, sheet_words['G9'].value.lower()],
                            [sheet_words.cell(row=i, column=1).value.lower() for i in range(2, sheet_words.max_row + 1)
                             if sheet_words.cell(row=i, column=1).value is not None]])
except:
    print("~~!Problem z plikiem kontrolnym")
    log_file.write("~~!Problem z plikiem kontrolnym")
#######################
try:
    book_links = load_workbook('links.xlsx')
######
    book_links_list = book_links.sheetnames
    sheet_links = book_words[book_words_list[0]]
######
    sheet_links = book_links.active
######
    webpages_to_research = [[sheet_links.cell(row=a, column=i).value for i in range(1, 11)] for a in range(2, sheet_links.max_row+1) if str(sheet_links.cell(row=a, column=5).value) > ""]
except:
    print("~~!Problem z bazą stron")
    log_file.write("~~!Problem z bazą stron")
##################################################################################################################

for webpage_district, webpage_city, webpage_type, webpage_url, css_path, method_url, filter_url, filter_title, base_url, webpage_approve in webpages_to_research:
    if webpage_approve == "TAK":
        print(">>> Odczyt", webpage_city)
        log_file.write(">>> Odczyt " + str(webpage_city) + "\n")
        obj__NbrOfArt = __NbrOfArt(method_url, webpage_url, css_path)
        len_art = obj__NbrOfArt.count_art()
        obj__DataOfArt = __DataOfArt(len_art, filter_title, filter_url, base_url)
        przetargi.extend(obj__DataOfArt.data_art())
        print("<<< Zakończono" + "\n")
        log_file.write("<<< Zakończono" + "\n")
log_file.close()

##################################################################################################################

for customer in range(len(book_words_list)):
    if filtr_slowo[customer][0][1] == 'tak':
        row = 0
        log_file = open('Logs\\Log_file_' + now_date + '_' + book_words_list[customer] + '.txt', 'a')
        log_file.write(str(now_date) + "\n")
        try:
            book_results = load_workbook('Results\\' + book_words_list[customer] + '_results.xlsx')
            book_results.create_sheet(now_date, 0)  # create a new sheet
            sheet_results = book_results.active  # get the reference to the active sheet
            # print("Znaleziono plik z wynikami dla", book_words_list[customer])
            log_file.write("Znaleziono plik z wynikami dla " + str(book_words_list[customer]) + "\n")
        except:
            book_results = Workbook()  # create a new workbook
            book_results.create_sheet(now_date, 0)  # create a new sheet
            sheet_results = book_results.active  # get the reference to the active sheet
            sheet_results.column_dimensions['A'].width = 7
            sheet_results.column_dimensions['B'].width = 100
            print("Utworzono nowy plik z wynikami dla", book_words_list[customer])
            log_file.write("Utworzono nowy plik z wynikami dla " + str(book_words_list[customer]) + "\n")
        ##################################################################################################################
        try:
            book_history = load_workbook('History\\' + book_words_list[customer] + '_history.xlsx')
            sheet_history = book_history.active
            history_list = [sheet_history.cell(row=i, column=1).value.lower() for i in range(1, sheet_history.max_row+1) if
                            sheet_history.cell(row=i, column=1).value is not None]
        except:
            book_history = Workbook()  # create a new workbook
            book_history.create_sheet("History", 0)
            sheet_history = book_history.active  # get the reference to the active sheet
            history_list = []
            print("Utworzono nowy plik z historia dla", book_words_list[customer])
            log_file.write("Utworzono nowy plik z historia dla " + str(book_words_list[customer]) + "\n")

        obj__ExcelHist = __ExcelHist()
        new_auctions = obj__ExcelHist.write_history()
        obj__ExcelArt = __ExcelArt(row, customer)
        obj__ExcelArt.write_excel()
        row = obj__ExcelArt.row
        # print("Dla klienta", book_words_list[customer], "Znaleziono nowych rekordow:", len(przetargi), ". W tym spelniających wymagania:", len(new_auctions))
        log_file.write("Dla klienta " + str(book_words_list[customer]) + " znaleziono nowych rekordow: " + str(len(przetargi)) + ". W tym spelniających wymagania: " + str(len(new_auctions)) + "\n")
        # print("")
        del obj__ExcelHist
        del obj__ExcelArt
        log_file.close()
################################### END ###########################################################
print("Zakończono raport z dnia ", now_date)
